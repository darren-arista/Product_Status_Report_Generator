import os
import tempfile
from datetime import datetime, timedelta
from flask import Flask, render_template, request, send_file, jsonify
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
import io
import sys

# For Vercel, use /tmp for writable storage
base_dir = '/tmp' if 'VERCEL' in os.environ else os.path.dirname(os.path.abspath(__file__))

# Create Flask app

app = Flask(__name__, 
    template_folder=os.path.dirname(os.path.abspath(__file__)),
    static_folder=None
)

# Disable debug mode in production
app.config.update(
    MAX_CONTENT_LENGTH=100 * 1024 * 1024,
    DEBUG=False
)

# ----------------- Enhanced Logo Handling -----------------
LOGO_PATHS = [
    os.path.join(base_dir, "Kith_logo.png"),
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "Kith_logo.png"),
]

# Add this function near the top of app.py
def init_vercel_environment():
    """Initialize Vercel-specific environment"""
    print(f"🔧 Vercel environment: {'VERCEL' in os.environ}")
    print(f"🔧 Python version: {sys.version}")
    print(f"🔧 Current directory: {os.getcwd()}")
    print(f"🔧 File directory: {os.path.dirname(os.path.abspath(__file__))}")
    
    # Ensure /tmp exists for logo
    if 'VERCEL' in os.environ:
        if not os.path.exists('/tmp'):
            os.makedirs('/tmp', exist_ok=True)
            print("✅ Created /tmp directory")
        
        # Copy logo to /tmp if it exists locally
        logo_source = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Kith_logo.png")
        if os.path.exists(logo_source):
            import shutil
            shutil.copy2(logo_source, '/tmp/Kith_logo.png')
            print(f"✅ Copied logo to /tmp")
    
    return True

# Call it right after Flask app initialization
app = Flask(__name__, template_folder=os.path.dirname(os.path.abspath(__file__)))
init_vercel_environment()

def find_logo_path():
    """Find the KITH logo in multiple possible locations"""
    print("🔍 Searching for KITH logo...")
    
    for path in LOGO_PATHS:
        try:
            absolute_path = os.path.abspath(path)
            exists = os.path.exists(absolute_path)
            
            if exists:
                file_size = os.path.getsize(absolute_path)
                print(f"✅ Found KITH logo at: {absolute_path} ({file_size} bytes)")
                return absolute_path
        except Exception as e:
            print(f"   ❌ Error checking path {path}: {e}")
            continue
    
    print("❌ KITH logo not found. Using text fallback.")
    return None

def create_fallback_logo():
    """Create a simple logo if none exists"""
    try:
        # For Vercel, create in /tmp
        logo_path = "/tmp/Kith_logo.png"
        
        from PIL import Image, ImageDraw, ImageFont
        
        # Create a simple logo
        img = Image.new('RGB', (200, 60), color='#1F4E78')
        d = ImageDraw.Draw(img)
        
        try:
            # Try to use a default font
            font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 28)
        except:
            font = ImageFont.load_default()
            
        d.text((60, 15), "KITH", fill=(255, 255, 255), font=font)
        img.save(logo_path)
        print(f"✅ Created fallback KITH logo at {logo_path}")
        return logo_path
    except Exception as e:
        print(f"❌ Could not create fallback logo: {e}")
        return None

# Initialize logo path
KITH_LOGO_PATH = find_logo_path()

# If no logo found, create one
if not KITH_LOGO_PATH:
    KITH_LOGO_PATH = create_fallback_logo()

# [Keep all your existing helper functions here - they should remain the same]
# [I've removed them for brevity but you should keep them in your actual file]

# ----------------- Helpers -----------------
def previous_business_day(date):
    d = date - timedelta(days=1)
    while d.weekday() >= 5:
        d -= timedelta(days=1)
    return d

def get_ordinal_suffix(day):
    """Return ordinal suffix for day (1st, 2nd, 3rd, etc.)"""
    if 11 <= day <= 13:
        return 'th'
    else:
        return {1: 'st', 2: 'nd', 3: 'rd'}.get(day % 10, 'th')

def format_date_for_header(date):
    """Format date as '5th Nov'25'"""
    day = date.day
    suffix = get_ordinal_suffix(day)
    return date.strftime(f"{day}{suffix} %b'%y")

def find_column(df_cols, candidates):
    lower_map = {str(c).strip().lower(): c for c in df_cols}
    for cand in candidates:
        cand_l = cand.strip().lower()
        if cand_l in lower_map:
            return lower_map[cand_l]
    return None

def read_file(file_stream, filename):
    # Debug the filename and extension
    print(f"Reading file: {filename}")
    print(f"File stream type: {type(file_stream)}")
    
    # Get the original filename from the FileStorage object if available
    original_filename = getattr(file_stream, 'filename', filename)
    print(f"Original filename: {original_filename}")
    
    ext = os.path.splitext(original_filename)[1].lower()
    print(f"Detected extension: {ext}")
    
    file_stream.seek(0)  # Reset file pointer
    
    try:
        if ext in [".xlsx", ".xls"]:
            print("Reading as Excel file...")
            # Use converters to ensure handle column is read as strings to preserve formatting
            return pd.read_excel(file_stream, dtype=str)
        elif ext == ".csv":
            print("Reading as CSV file...")
            # Try different encodings
            try:
                return pd.read_csv(file_stream, dtype=str, encoding='utf-8')
            except UnicodeDecodeError:
                try:
                    file_stream.seek(0)
                    return pd.read_csv(file_stream, dtype=str, encoding='latin-1')
                except UnicodeDecodeError:
                    file_stream.seek(0)
                    return pd.read_csv(file_stream, dtype=str, encoding='cp1252')
        else:
            # If no extension detected, try to detect file type by content
            print("No extension detected, trying to detect file type...")
            file_stream.seek(0)
            first_bytes = file_stream.read(100)
            file_stream.seek(0)
            
            # Check if it's a CSV by looking for commas
            if b',' in first_bytes:
                print("Detected CSV by content (commas found)")
                try:
                    return pd.read_csv(file_stream, dtype=str, encoding='utf-8')
                except UnicodeDecodeError:
                    file_stream.seek(0)
                    return pd.read_csv(file_stream, dtype=str, encoding='latin-1')
            # Check if it's Excel by signature
            elif first_bytes.startswith(b'PK'):  # Excel files are ZIP archives
                print("Detected Excel by content (ZIP signature)")
                return pd.read_excel(file_stream, dtype=str)
            else:
                raise ValueError(f"Unsupported file type. Extension: '{ext}', File: {original_filename}")
    except Exception as e:
        print(f"Error reading file {original_filename}: {str(e)}")
        print(f"File extension was: '{ext}'")
        raise

def apply_light_borders(ws, start_row=1, start_col=1, end_row=None, end_col=None):
    """Apply light borders to maintain gridline appearance"""
    if end_row is None:
        end_row = ws.max_row
    if end_col is None:
        end_col = ws.max_column
    
    # Define light border style (similar to Excel's default gridlines)
    light_border = Border(
        left=Side(style='thin', color='FFD0D0D0'),
        right=Side(style='thin', color='FFD0D0D0'),
        top=Side(style='thin', color='FFD0D0D0'),
        bottom=Side(style='thin', color='FFD0D0D0')
    )
    
    # Apply borders to all cells in range
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, 
                           min_col=start_col, max_col=end_col):
        for cell in row:
            cell.border = light_border

def apply_black_borders(ws, start_row, start_col, end_row, end_col):
    """Apply black borders to cells in summary sheet"""
    # Define black border style
    black_border = Border(
        left=Side(style='thin', color='FF000000'),
        right=Side(style='thin', color='FF000000'),
        top=Side(style='thin', color='FF000000'),
        bottom=Side(style='thin', color='FF000000')
    )
    
    # Apply borders to all cells in range
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            # Check if this is not a MergedCell by trying to access its value
            try:
                _ = cell.value
                cell.border = black_border
            except AttributeError:
                # This is a MergedCell, skip it
                continue

# ----------------- Core report + summary creation -----------------
def create_report_and_summary(today_file, yesterday_raw_file, yesterday_final_file, output_filename):
    try:
        # --- Read and clean data ---
        print("Reading today's file...")
        df_today_raw = read_file(today_file, "today")
        print(f"Today's file shape: {df_today_raw.shape}")
        print("Today's file columns:", df_today_raw.columns.tolist())
        
        print("Reading yesterday's raw file...")
        df_yest_raw = read_file(yesterday_raw_file, "yesterday")
        print(f"Yesterday's file shape: {df_yest_raw.shape}")
        print("Yesterday's file columns:", df_yest_raw.columns.tolist())
        
        df_today_raw.columns = [str(c).strip() for c in df_today_raw.columns]
        df_yest_raw.columns = [str(c).strip() for c in df_yest_raw.columns]

        # Print available columns for debugging
        print("Available columns in today's file:", df_today_raw.columns.tolist())
        print("Available columns in yesterday's file:", df_yest_raw.columns.tolist())

        handle_col = find_column(df_today_raw.columns, ["Handle", "Handles", "handle"])
        title_col = find_column(df_today_raw.columns, ["Title", "title", "Name"])
        id_col = find_column(df_today_raw.columns, ["ID", "Id", "Product ID", "ProductID"])
        vendor_col = find_column(df_today_raw.columns, ["Vendor", "vendor", "Manufacturer"])
        type_col = find_column(df_today_raw.columns, ["Type", "type", "Custom Product Type", "custom product type"])
        tags_col = find_column(df_today_raw.columns, ["Tags", "tags"])
        published_col_today = find_column(df_today_raw.columns, ["Published", "published", "Is Published", "is_published"])
        
        # Debug column findings
        print("Found columns - Handle:", handle_col, "Title:", title_col, "ID:", id_col)
        print("Vendor:", vendor_col, "Type:", type_col, "Tags:", tags_col, "Published:", published_col_today)
        
        for col, name in [(handle_col, "Handle"), (title_col, "Title"), (id_col, "ID"),
                          (vendor_col, "Vendor"), (type_col, "Type"),
                          (tags_col, "Tags"), (published_col_today, "Published")]:
            if col is None:
                error_msg = f"Today's file is missing required column: {name}. Available columns: {df_today_raw.columns.tolist()}"
                print(error_msg)
                raise ValueError(error_msg)

        handle_col_y = find_column(df_yest_raw.columns, ["Handle", "Handles", "handle"])
        published_col_y = find_column(df_yest_raw.columns, ["Published", "published", "Is Published", "is_published"])
        
        print("Yesterday's columns - Handle:", handle_col_y, "Published:", published_col_y)
        
        if handle_col_y is None or published_col_y is None:
            error_msg = f"Yesterday's raw file must contain 'Handle' and 'Published' columns. Available: {df_yest_raw.columns.tolist()}"
            print(error_msg)
            raise ValueError(error_msg)

        today_date = datetime.today().date()
        yesterday_date = previous_business_day(today_date)

        # Format dates for column headers
        today_header = f"Published Status on {format_date_for_header(today_date)}"
        yesterday_header = f"Published Status on {format_date_for_header(yesterday_date)}"
        
        print(f"Today's header: {today_header}")
        print(f"Yesterday's header: {yesterday_header}")

        # --- Vectorized processing ---
        df_today = df_today_raw.copy()
        df_today["TodayStatus"] = df_today[published_col_today].str.upper().isin(["TRUE", "T", "1", "YES", "Y"])
        df_yest = df_yest_raw.copy()
        df_yest["YesterdayStatus"] = df_yest[published_col_y].str.upper().isin(["TRUE", "T", "1", "YES", "Y"])

        y_map = df_yest.set_index(handle_col_y)["YesterdayStatus"].to_dict()
        
        # Ensure handles are treated as strings and properly formatted
        df_today[handle_col] = df_today[handle_col].astype(str).str.strip()
        
        # Simple solution: Show "N/A" for new products, keep original logic for calculations
        df_today["YesterdayStatus"] = df_today[handle_col].map(y_map).fillna(False)
        df_today["TodayStr"] = df_today["TodayStatus"].map({True: "TRUE", False: "FALSE"})
        
        # Create display version with "N/A" for new products
        df_today["YestStr"] = df_today[handle_col].map(y_map)
        df_today["YestStr"] = df_today["YestStr"].map({True: "TRUE", False: "FALSE"})
        df_today["YestStr"] = df_today["YestStr"].fillna("N/A")

        df_final = df_today[[handle_col, title_col, id_col, vendor_col, type_col, tags_col, "TodayStr", "YestStr"]].copy()
        df_final.columns = [
            "Handle", "Title", "Product ID's", "Vendor", "Custom Product Type", "Tags",
            today_header, yesterday_header
        ]

        # Hidden products calculation uses the original boolean logic (no N/A involved)
        df_hidden = df_final[(df_today["YesterdayStatus"]) & (~df_today["TodayStatus"])]
        df_other = df_final.drop(df_hidden.index)
        df_ordered = pd.concat([df_hidden, df_other], ignore_index=True)

        # Convert Product ID's to numeric (keeping as numbers)
        df_ordered["Product ID's"] = pd.to_numeric(df_ordered["Product ID's"], errors='coerce')
        
        # Keep handles as original strings - no modifications
        out_rows = df_ordered.astype(str).values.tolist()

        # --- Workbook setup ---
        wb = Workbook()

        # Create sheet with today's date as name in format: 7th Nov'25
        sheet_name = format_date_for_header(today_date)
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
            wb.remove(wb["Sheet"])
        ws = wb.create_sheet(sheet_name)

        # Enable gridlines for the sheet
        ws.sheet_view.showGridLines = True

        # Header
        ws.append(list(df_final.columns))
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"

        # Data - write rows manually to ensure proper formatting
        for row_data in out_rows:
            ws.append(row_data)

        # Apply proper formatting to each column
        handle_col_idx = df_final.columns.get_loc("Handle") + 1
        prod_id_col_idx = df_final.columns.get_loc("Product ID's") + 1
        title_col_idx = df_final.columns.get_loc("Title") + 1
        
        for row in ws.iter_rows(min_row=2, max_row=len(out_rows)+1, min_col=1, max_col=len(df_final.columns)):
            for c_idx, cell in enumerate(row, start=1):
                # Handle column: Set to General format - ALLOW SCIENTIFIC NOTATION IN CELL DISPLAY
                if c_idx == handle_col_idx:
                    # General format - let Excel auto-detect and display scientific notation if needed
                    cell.number_format = 'General'
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    # REMOVED: The conversion that was preventing scientific notation
                    # The original string value will be preserved, allowing Excel to display scientific notation
                    # while keeping the full number in the formula bar
                
                # Product ID's column: set as number format (no decimals)
                elif c_idx == prod_id_col_idx:
                    cell.number_format = '0'  # Number format with no decimals
                    # Convert back to number if it's a numeric string
                    if cell.value and str(cell.value).replace('.', '').replace('-', '').isdigit():
                        try:
                            cell.value = int(float(cell.value))
                        except (ValueError, TypeError):
                            pass  # Keep as string if conversion fails
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Title column gets left alignment
                elif c_idx == title_col_idx:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # Other columns get center alignment
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        # Apply light borders to maintain gridline appearance even with conditional formatting
        apply_light_borders(ws, start_row=1, end_row=len(out_rows)+1, end_col=len(df_final.columns))

        # Conditional formatting Product Status Report
        green_fill = PatternFill(start_color="FFC6E0BA", end_color="FFC6E0BA", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFFFACD", end_color="FFFFFACD", fill_type="solid")
        today_col_letter = get_column_letter(len(df_final.columns) - 1)
        yest_col_letter = get_column_letter(len(df_final.columns))
        rule_hidden = FormulaRule(formula=[f'AND(${today_col_letter}2="FALSE", ${yest_col_letter}2="TRUE")'], fill=green_fill)
        rule_other = FormulaRule(formula=[f'NOT(AND(${today_col_letter}2="FALSE", ${yest_col_letter}2="TRUE"))'], fill=yellow_fill)
        ws.conditional_formatting.add(f"A2:{get_column_letter(len(df_final.columns))}{len(out_rows)+1}", rule_hidden)
        ws.conditional_formatting.add(f"A2:{get_column_letter(len(df_final.columns))}{len(out_rows)+1}", rule_other)

        # Column widths - updated as requested
        width_map = {
            "Handle": 27.57,
            "Title": 49.29,
            "Product ID's": 19,
            "Vendor": 19,
            "Custom Product Type": 19,
            "Tags": 19,
            today_header: 19,
            yesterday_header: 19
        }
        for c_idx, header in enumerate(df_final.columns, start=1):
            ws.column_dimensions[get_column_letter(c_idx)].width = width_map.get(header, 19)

        # --- Summary sheet ---
        def init_summary_sheet(wb_obj):
            if "Summary" in wb_obj.sheetnames:
                wb_obj.remove(wb_obj["Summary"])
            ws_sum = wb_obj.create_sheet("Summary", 0)
            
            # Enable gridlines for the summary sheet
            ws_sum.sheet_view.showGridLines = True
            
            # Set column widths
            for col in ["A", "B", "C", "D", "E"]:
                ws_sum.column_dimensions[col].width = 37
            
            # Add KITH logo - SMALL and CENTERED in column C
            ws_sum.merge_cells("A1:E1")
            ws_sum.row_dimensions[1].height = 60  # Reduced height for smaller logo
            
            # Try to add the logo - SMALL and CENTERED in column C
            logo_added = False
            if KITH_LOGO_PATH:
                try:
                    img = XLImage(KITH_LOGO_PATH)
                    # EXACT logo size as requested
                    img.width = 96  # 1 inch = 96 pixels (at 96 DPI)
                    img.height = 52  # 0.54 inch = 52 pixels (at 96 DPI)
                    
                    # Place it in column C (middle of A-E) - CENTERED
                    ws_sum.add_image(img, "C1")
                    logo_added = True
                    print("✅ KITH logo added successfully with exact dimensions (1\" x 0.54\") and centered in column C")
                except Exception as e:
                    print(f"❌ Error adding KITH logo from {KITH_LOGO_PATH}: {e}")
            
            # If logo couldn't be added, show text instead - CENTERED in the merged cells A1:E1
            if not logo_added:
                # Set the value in A1 (top-left cell of the merge)
                ws_sum["A1"] = "KITH"
                ws_sum["A1"].font = Font(size=18, bold=True, color="1F4E78")
                ws_sum["A1"].alignment = Alignment(horizontal="center", vertical="center")
                print("📝 Using small text 'KITH' instead of logo (centered in merged cells A1:E1)")

            month_name = datetime.today().strftime("%B")
            ws_sum["A2"] = month_name
            ws_sum["A2"].font = Font(bold=True, size=14, color="FFFFFFFF")  # White font color
            ws_sum["A2"].alignment = Alignment(horizontal="center", vertical="center")
            ws_sum.merge_cells("B2:E2")
            ws_sum["B2"] = "Product Overview"
            ws_sum["B2"].font = Font(bold=True, size=14, color="FFFFFFFF")  # White font color
            ws_sum["B2"].alignment = Alignment(horizontal="center", vertical="center")
            ws_sum.row_dimensions[2].height = 32  # Row height 32
            
            # Blue header background
            blue_fill = PatternFill(start_color="FFA5A5A5", end_color="FFA5A5A5", fill_type="solid")
            for col in range(1, 6):
                cell = ws_sum.cell(row=2, column=col)
                cell.fill = blue_fill
                cell.font = Font(bold=True, size=14, color="FFFFFFFF")  # White font color
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Sub headers - ORANGE color
            sub_headers = ["Date", "Published Products", "Unpublished Products", "Total Products", "Hidden Products Status"]
            for i, sh in enumerate(sub_headers, start=1):
                cell = ws_sum.cell(row=3, column=i, value=sh)
                cell.font = Font(bold=True, color="FFFFA500", size=12)  # Orange color
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Set row height for third row
            ws_sum.row_dimensions[3].height = 28  # Row height 28
            
            return ws_sum

        ws_summary = init_summary_sheet(wb)

        published_count = int(df_today["TodayStatus"].sum())
        unpublished_count = int((~df_today["TodayStatus"]).sum())
        # Use the original boolean logic for hidden count (no N/A involved)
        hidden_count = int(((~df_today["TodayStatus"]) & (df_today["YesterdayStatus"])).sum())
        total_products = len(df_today)

        # Handle yesterday's final file if provided - append today's data below yesterday's
        historical_rows = []
        if yesterday_final_file:
            try:
                print("Reading yesterday's final report...")
                yesterday_final_file.seek(0)  # Reset file pointer
                wb_y = load_workbook(yesterday_final_file, data_only=True, read_only=True)
                if "Summary" in wb_y.sheetnames:
                    ws_y = wb_y["Summary"]
                    for row in ws_y.iter_rows(min_row=4, values_only=True):
                        if row[0] is not None:
                            historical_rows.append(row)
                    print(f"Found {len(historical_rows)} historical rows from yesterday's final report")
            except Exception as e:
                print(f"Warning: Could not read yesterday's final file: {e}")

        # Add historical rows first (yesterday's data)
        for row in historical_rows:
            ws_summary.append(row)

        # Insert today's summary row after historical data
        date_str = datetime.today().strftime("%m/%d/%Y")
        # Use "-" when no handles are hidden and keep cell white (no formatting)
        hidden_text = "-" if hidden_count == 0 else f"{hidden_count} handles were Hidden today."
        vals = [date_str, published_count, unpublished_count, total_products, hidden_text]
        ws_summary.append(vals)

        # Apply BLACK borders to summary sheet cells (rows 4 to max_row, columns 1 to 5)
        # Skip rows 1-3 which have merged cells
        apply_black_borders(ws_summary, start_row=4, start_col=1, end_row=ws_summary.max_row, end_col=5)

        # Highlight all rows and center align - updated formatting
        fill_yellow = PatternFill(start_color="FFFFFFCC", end_color="FFFFFFCC", fill_type="solid")
        fill_green = PatternFill(start_color="FFC6E0B4", end_color="FFC6E0B4", fill_type="solid")
        
        for row_idx, row in enumerate(ws_summary.iter_rows(min_row=4, max_row=ws_summary.max_row, min_col=1, max_col=5), start=4):
            for cell in row:
                # Use try/except to handle MergedCell objects
                try:
                    # Try to access value to check if it's a regular cell
                    current_value = cell.value
                    
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Set number format for numeric columns (B, C, D)
                    if cell.column_letter in ['B', 'C', 'D']:  # Published Products, Unpublished Products, Total Products
                        cell.number_format = '0'  # Number format with no decimals
                        # Ensure the value is stored as a number
                        if cell.value is not None and str(cell.value).isdigit():
                            try:
                                cell.value = int(cell.value)
                            except (ValueError, TypeError):
                                pass  # Keep original value if conversion fails
                    
                    # Apply yellow fill to all columns except Hidden Products Status (column E)
                    if cell.column_letter != 'E':  # Columns A, B, C, D get yellow
                        cell.fill = fill_yellow
                    else:  # Column E (Hidden Products Status) - check individual cell value
                        # Apply green ONLY if the cell contains text about hidden handles (not "-")
                        if cell.value != "-" and cell.value is not None and "hidden" in str(cell.value).lower():
                            cell.fill = fill_green
                        else:
                            # Keep white (no fill) when value is "-" (no hidden handles)
                            cell.fill = PatternFill()  # No fill
                except AttributeError:
                    # This is a MergedCell, skip it (shouldn't happen from row 4)
                    continue
            
            # Format Date column - use try/except
            date_cell = row[0]
            try:
                # Check if it's a regular cell
                current_value = date_cell.value
                
                if isinstance(date_cell.value, str):
                    try:
                        date_obj = datetime.strptime(date_cell.value, "%m/%d/%Y")
                    except Exception:
                        date_obj = datetime.today()
                    date_cell.value = date_obj
                date_cell.number_format = 'MM/DD/YYYY'
            except AttributeError:
                # Skip MergedCell (shouldn't happen from row 4)
                pass

        # Save to bytes buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        print("Report generation completed successfully")
        return output
        
    except Exception as e:
        print(f"Error in create_report_and_summary: {str(e)}")
        import traceback
        traceback.print_exc()
        raise

# ----------------- Routes -----------------
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/generate', methods=['POST'])
def generate_report():
    try:
        print("Generate route called")  # Debug log
        print(f"Request content length: {request.content_length}")
        print(f"Request files: {list(request.files.keys())}")
        
        # Check required files
        if 'today_file' not in request.files or 'yesterday_raw_file' not in request.files:
            return jsonify({'error': 'Today\'s file and yesterday\'s raw file are required'}), 400
        
        today_file = request.files['today_file']
        yesterday_raw_file = request.files['yesterday_raw_file']
        yesterday_final_file = request.files.get('yesterday_final_file')
        
        print(f"Files received - Today: {today_file.filename}, Yesterday Raw: {yesterday_raw_file.filename}")
        
        if today_file.filename == '' or yesterday_raw_file.filename == '':
            return jsonify({'error': 'No file selected'}), 400

        # Check file sizes
        today_file.seek(0, 2)  # Seek to end
        today_size = today_file.tell()
        today_file.seek(0)  # Reset to beginning
        
        yesterday_raw_file.seek(0, 2)
        yesterday_size = yesterday_raw_file.tell()
        yesterday_raw_file.seek(0)
        
        print(f"File sizes - Today: {today_size} bytes, Yesterday: {yesterday_size} bytes")

        # Generate output filename
        output_filename = request.form.get('output_filename', 'product_status_report.xlsx')
        if not output_filename.endswith('.xlsx'):
            output_filename += '.xlsx'

        print(f"Generating report: {output_filename}")

        # Create the report
        output_file = create_report_and_summary(
            today_file,
            yesterday_raw_file,
            yesterday_final_file,
            output_filename
        )

        print("Report generated successfully, sending file...")
        return send_file(
            output_file,
            as_attachment=True,
            download_name=output_filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        print(f"Error in generate_report route: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# Add error handler for 413
@app.errorhandler(413)
def too_large(e):
    return jsonify({'error': 'File too large. Maximum file size is 100MB. Please use smaller files or split your data.'}), 413

# This is important for Vercel
if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))
else:
    application = app